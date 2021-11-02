import os
from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup as BS
import pyexcel
import pandas as pd

CURR_FOLDER = os.path.abspath('')


class ExcelReader:
    def __init__(self):
        allFiles = os.listdir(CURR_FOLDER + "/" + "downloads")
        self.validFiles = self.filterFiles(allFiles)
        self.convertXlsToXlsx()

    def convertXlsToXlsx(self):
        for file in self.validFiles:
            if file.split(".")[-1] == "xls":
                pyexcel.save_book_as(file_name=CURR_FOLDER + '/downloads/' + file,
                                     dest_file_name=CURR_FOLDER + '/downloads/' + file.split(".")[0] + '.xlsx')
                os.remove(CURR_FOLDER + '/downloads/' + file)

    def filterFiles(self, files):
        valid = []
        for file in files:
            fileP = file.split(".")[-1]
            if fileP == "xls" or fileP == "xlsx":
                valid.append(file)
        return valid


    def getSheet(self, wb):
        for sheet in wb.sheetnames:
            if sheet == "РК":
                return wb["РК"]
            elif sheet == "Респуб":
                return wb["Респуб"]
            elif sheet == "по РК":
                return wb["по РК"]
        return 0

    def getTypeDoc(self, sheet, index):
        value = sheet.cell(index, 4)
        if value.value == "РБ" or value.value == "Республиканский бюджет":
            return 1

        elif value.value == "Всего поступило в Гос.бюджет":
            return 2

        elif value.value == "Факт":
            return 3

        return 0

    def deleteDownlFiles(self):
        for file in self.validFiles:
            if file.split(".")[-1] == "xls":
                file = file + "x"

            os.remove(CURR_FOLDER + "/" + "downloads" + "/" + file)

    def getColm(self, typedoc):
        if typedoc == 1:
            return 3, 4, 5
        elif typedoc == 2:
            return 4, 6, 8
        elif typedoc == 3:
            return 4, 7, 10

    def read(self):
        data = []
        for file in self.validFiles:
            if file.split(".")[-1] == "xls":
                file = file + "x"
            wb = load_workbook(CURR_FOLDER + "/" + "downloads" + "/" + file)
            sheet = self.getSheet(wb)
            if sheet == 0:
                continue
            index = 1
            nextValidData = False
            typedoc = 0
            while index <= sheet.max_row:
                code = sheet.cell(index, 1)

                if typedoc == 0:
                    typedoc = self.getTypeDoc(sheet, index)
                else:
                    gb, rb, mb = self.getColm(typedoc)
                    month, year = getMonthAndYear(file)

                    if code.value == 1:
                        nextValidData = True

                    if code.value is None:
                        index += 1
                        continue

                    if sheet.cell(index, 2).value == 2:
                        index += 1
                        continue

                    if nextValidData:
                        data.append(
                                {
                                    "ГОД": year,
                                    "МЕСЯЦ": month,
                                    "КБК": code.value,
                                    "ГБ": sheet.cell(index, gb).value,
                                    "РБ": sheet.cell(index, rb).value,
                                    "МБ": sheet.cell(index, mb).value
                                }
                            )

                index += 1
            wb.close()
        return data


class Downloader:
    def __init__(self):
        self.URL = "https://kgd.gov.kz/ru/content/fakticheskie-postupleniya-po-nalogam-i-platezham-v-gosudarstvennyy-byudzhet-za-2002-2018-gg"
        self.soup = self.getSoup(requests.get(self.URL).text)

    def downloadXlsx(self):
        urlsToXlsx = self.getLinks()
        if not os.path.isdir(CURR_FOLDER + "/" + "downloads"):
            os.mkdir(CURR_FOLDER + "/" + "downloads")

        for url in urlsToXlsx:
            saveXlsx(url)

    def getLinks(self):
        links = []
        table = self.soup.find("tbody")
        trs = table.find_all("tr")
        del trs[0]
        currYear = 2021
        for tr in trs:
            tds = tr.find_all("td", class_="rtecenter")
            del tds[0]
            currMonth = 1
            for td in tds:
                aTag = td.find("a")
                if aTag is None:
                    month_, year_ = getMonthAndYear(str(currMonth) + "-" + str(currYear) + ".xlsx")
                    print(f"file {str(month_) + '-' + year_} missing for download")
                    currMonth += 1
                    continue

                url = self.addSitePrefix(aTag.get("href"))
                links.append(
                    {
                        "currYear": currYear,
                        "currMonth": currMonth,
                        "url": url
                    }
                )
                currMonth += 1
            currYear -= 1
        return links

    def addSitePrefix(self, url):
        if "https://" in url:
            return url
        else:
            return "https://kgd.gov.kz/" + url

    def getSoup(self, html):
        return BS(html, "lxml")


def getMonthAndYear(file):
    monthIndex = file.split("-")[0]
    yearIndex = file.split("-")[-1].split(".")[0]
    monthIndex = int(monthIndex)
    if monthIndex == 1:
        month = "январь"

    elif monthIndex == 2:
        month = "февраль"

    elif monthIndex == 3:
        month = "март"

    elif monthIndex == 4:
        month = "апрель"

    elif monthIndex == 5:
        month = "май"

    elif monthIndex == 6:
        month = "июнь"

    elif monthIndex == 7:
        month = "июль"

    elif monthIndex == 8:
        month = "август"

    elif monthIndex == 9:
        month = "сентябрь"

    elif monthIndex == 10:
        month = "октябрь"

    elif monthIndex == 11:
        month = "ноябрь"

    elif monthIndex == 12:
        month = "декабрь"

    else:
        month = ""

    return month, yearIndex


def saveXlsx(url):
    extension = url["url"].split(".")[-1]
    with open(CURR_FOLDER + "/" + "downloads" + "/" + str(url["currMonth"]) + "-" + str(
            url["currYear"]) + "." + extension, "wb") as file:
        r = requests.get(url["url"])
        file.write(r.content)


def main():
    print("I'm starting to download files")
    downloader = Downloader()
    downloader.downloadXlsx()
    print("File upload completed")
    print("Converting .xls files to .xlsx")
    reader = ExcelReader()
    print("I start collecting data from files")
    data = reader.read()
    reader.deleteDownlFiles()
    df = pd.DataFrame(data)
    print(data)
    print(df)

if __name__ == '__main__':
    main()
